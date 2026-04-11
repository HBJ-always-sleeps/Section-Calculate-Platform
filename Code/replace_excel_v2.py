# -*- coding: utf-8 -*-
"""
数据替换脚本 - 简化版
"""
import pandas as pd
import openpyxl
from datetime import datetime
import os

os.chdir(r"D:\断面算量平台\测试文件")

target_file = "2026年2月月进度工程量表（提交版）.xlsx"
source_file = "内湾段分层图（全航道）_分类汇总_20260316_154717.xlsx"

print("Step 1: Reading source file...")
source_design = pd.read_excel(source_file, sheet_name='设计量汇总')
source_overbreak = pd.read_excel(source_file, sheet_name='超挖汇总')
print(f"Source columns: {list(source_design.columns)}")

# Sheet mapping: 目标文件sheet名 -> 源文件列名
# 注意：目标文件用"粘土"，源文件用"黏土"（不同的汉字！）
sheet_map = {
    '1级淤泥': '1级淤泥',
    '2级淤泥': '2级淤泥', 
    '3级淤泥': '3级淤泥',
    '1级填土': '1级填土',
    '5级粘土': '5级黏土',  # 目标文件: "粘土"，源文件: "黏土"
}

COEFF = 0.6

print("\nStep 2: Opening target file...")
wb = openpyxl.load_workbook(target_file)
print(f"Sheets: {wb.sheetnames}")

for target_sheet, source_col in sheet_map.items():
    # Find matching sheet
    actual_name = None
    for sn in wb.sheetnames:
        if sn.strip() == target_sheet.strip():
            actual_name = sn
            break
    
    if not actual_name:
        print(f"Skip: {target_sheet}")
        continue
    
    print(f"\nProcessing: {actual_name}")
    ws = wb[actual_name]
    
    # Build pile data
    pile_data = {}
    for idx, row in source_design.iterrows():
        pile = str(row['桩号']).strip()
        design_val = row.get(source_col, 0) or 0
        over_val = source_overbreak.loc[idx, source_col] if source_col in source_overbreak.columns else 0
        over_val = over_val or 0
        
        try:
            design_val = float(design_val)
            over_val = float(over_val)
        except:
            design_val = 0
            over_val = 0
        
        pile_data[pile] = {
            'design': round(design_val * COEFF, 3),
            'over': round(over_val * COEFF, 3)
        }
    
    print(f"  Piles loaded: {len(pile_data)}")
    
    # Process rows (starting row 10, every 2 rows)
    updated = 0
    for r in range(10, ws.max_row + 1, 2):
        pile_cell = ws.cell(row=r, column=2).value
        if not pile_cell:
            continue
            
        pile = str(pile_cell).strip()
        if pile not in pile_data:
            continue
        
        data = pile_data[pile]
        
        # Step 1: Copy current to previous (L->H, M->I, N->J, O->K)
        for i, (src, dst) in enumerate([(12, 8), (13, 9), (14, 10), (15, 11)]):
            val = ws.cell(row=r, column=src).value
            if val is not None:
                ws.cell(row=r, column=dst).value = val
        
        # Step 2: Set new current values
        ws.cell(row=r, column=12).value = data['design']  # L: 开挖面积
        ws.cell(row=r, column=14).value = data['over']    # N: 超挖面积
        updated += 1
    
    print(f"  Updated: {updated}")
    
    # Step 3: Calculate volumes (skip merged cells and row 499)
    merged_ranges = list(ws.merged_cells.ranges)
    
    def is_merged(row, col):
        for merged_range in merged_ranges:
            if merged_range.min_row <= row <= merged_range.max_row and merged_range.min_col <= col <= merged_range.max_col:
                return True
        return False
    
    for r in range(10, ws.max_row, 2):
        curr_design = ws.cell(row=r, column=12).value or 0
        curr_over = ws.cell(row=r, column=14).value or 0
        next_design = ws.cell(row=r+2, column=12).value or 0
        next_over = ws.cell(row=r+2, column=14).value or 0
        
        try:
            vol_design = (float(curr_design) + float(next_design)) / 2 * 25
            vol_over = (float(curr_over) + float(next_over)) / 2 * 25
        except:
            vol_design = 0
            vol_over = 0
        
        # Only write if not merged and not row 499
        if not is_merged(r+1, 13) and r+1 != 499:
            ws.cell(row=r+1, column=13).value = round(vol_design, 2)
        if not is_merged(r+1, 15) and r+1 != 499:
            ws.cell(row=r+1, column=15).value = round(vol_over, 2)
    
    # Step 4: Handle row 500 (合计行) - copy current totals to previous totals
    # L500->H500, M500->I500, N500->J500, O500->K500
    for src, dst in [(12, 8), (13, 9), (14, 10), (15, 11)]:
        val = ws.cell(row=500, column=src).value
        if val is not None:
            ws.cell(row=500, column=dst).value = val
    print(f"  Row 500 totals copied")

# Save
output = f"2026年2月月进度工程量表_替换后_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
wb.save(output)
wb.close()
print(f"\nDone! Output: {output}")
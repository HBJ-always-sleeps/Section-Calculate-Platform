# -*- coding: utf-8 -*-
"""
数据替换脚本：
1. 将目标文件的前5个sheet的本期剩余工程量覆盖上期剩余工程量
2. 用源文件的数据乘0.6系数填入本期剩余工程量的面积
3. 计算体积：前后两个面积之和除以二再乘25
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import shutil
from datetime import datetime

# 文件路径
target_file = r"D:\断面算量平台\测试文件\2026年2月月进度工程量表（提交版）.xlsx"
source_file = r"D:\断面算量平台\测试文件\内湾段分层图（全航道）_分类汇总_20260316_154717.xlsx"

# 创建备份
backup_file = target_file.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
shutil.copy(target_file, backup_file)
print(f"已创建备份: {backup_file}")

# 要处理的工作表映射：目标文件sheet名 -> 源文件列名
# 注意：目标文件用"粘土"，源文件用"黏土"
sheet_mapping = {
    '1级淤泥': '1级淤泥',
    '2级淤泥': '2级淤泥',
    '3级淤泥': '3级淤泥',
    '1级填土': '1级填土',
    '5级粘土': '5级黏土',  # 目标文件用"粘土"，源文件用"黏土"
}

# 系数
COEFFICIENT = 0.6

def get_cell_value(ws, row, col):
    """安全获取单元格值，处理合并单元格"""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        # 找到合并区域的主单元格
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return None
    return cell.value

def set_cell_value(ws, row, col, value):
    """安全设置单元格值，处理合并单元格"""
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        # 找到合并区域的主单元格
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(row=merged_range.min_row, column=merged_range.min_col).value = value
                return
    cell.value = value

# 读取源文件数据
print("\n正在读取源文件...")
source_design = pd.read_excel(source_file, sheet_name='设计量汇总', header=0)
source_overbreak = pd.read_excel(source_file, sheet_name='超挖汇总', header=0)

# 显示源文件列名以便调试
print(f"设计量汇总列名: {list(source_design.columns)}")
print(f"超挖汇总列名: {list(source_overbreak.columns)}")

# 打开目标文件
print("\n正在处理目标文件...")
wb = openpyxl.load_workbook(target_file)

# 显示所有工作表名
print(f"目标文件工作表: {wb.sheetnames}")

# 处理每个工作表
for target_sheet_name, source_col_name in sheet_mapping.items():
    # 找到实际的工作表名（处理空格问题）
    actual_sheet_name = None
    for sn in wb.sheetnames:
        if sn.strip() == target_sheet_name.strip():
            actual_sheet_name = sn
            break
    
    if actual_sheet_name is None:
        print(f"警告: 未找到工作表 '{target_sheet_name}'，跳过")
        continue
    
    print(f"\n处理工作表: {actual_sheet_name}")
    ws = wb[actual_sheet_name]
    
    # 收集需要更新的桩号数据
    pile_data = {}  # {桩号: {design_area, overbreak_area}}
    
    # 遍历源文件数据，收集桩号和对应数据
    for idx, row in source_design.iterrows():
        pile_name = str(row['桩号']).strip()
        design_value = row.get(source_col_name, 0)
        overbreak_value = source_overbreak.loc[idx, source_col_name] if source_col_name in source_overbreak.columns else 0
        
        # 转换为数值
        try:
            design_value = float(design_value) if pd.notna(design_value) else 0
        except:
            design_value = 0
        try:
            overbreak_value = float(overbreak_value) if pd.notna(overbreak_value) else 0
        except:
            overbreak_value = 0
        
        # 乘以系数
        design_area = design_value * COEFFICIENT
        overbreak_area = overbreak_value * COEFFICIENT
        
        pile_data[pile_name] = {
            'design_area': round(design_area, 3),
            'overbreak_area': round(overbreak_area, 3)
        }
    
    print(f"  从源文件读取到 {len(pile_data)} 个桩号数据")
    
    # 遍历目标文件，找到桩号行并更新
    # 目标文件结构：从第10行开始，每个断面占2行
    # 奇数行(10,12,14...)是面积行，偶数行(11,13,15...)是体积行
    # 
    # 列对应：
    # H列(8): 上期剩余开挖面积
    # I列(9): 上期剩余开挖体积
    # J列(10): 上期剩余超挖面积
    # K列(11): 上期剩余超挖体积
    # L列(12): 本期剩余开挖面积
    # M列(13): 本期剩余开挖体积
    # N列(14): 本期剩余超挖面积
    # O列(15): 本期剩余超挖体积
    
    updated_count = 0
    
    # 第一步：将本期剩余工程量复制到上期剩余工程量
    print(f"  步骤1: 复制本期剩余到上期剩余...")
    for excel_row in range(10, ws.max_row + 1, 2):  # 面积行
        for col_offset in range(4):  # 4列数据：面积和体积
            src_col = 12 + col_offset  # L-O列
            dst_col = 8 + col_offset   # H-K列
            src_val = get_cell_value(ws, excel_row, src_col)
            if src_val is not None:
                set_cell_value(ws, excel_row, dst_col, src_val)
    
    # 第二步：填入新的本期剩余工程量面积（源数据 * 0.6）
    print(f"  步骤2: 填入新的本期剩余面积...")
    for excel_row in range(10, ws.max_row + 1, 2):  # 面积行
        pile_cell = get_cell_value(ws, excel_row, 2)  # B列是桩号
        if pile_cell:
            pile_name = str(pile_cell).strip()
            if pile_name in pile_data:
                data = pile_data[pile_name]
                # L列(12)：开挖面积，N列(14)：超挖面积
                set_cell_value(ws, excel_row, 12, data['design_area'])
                set_cell_value(ws, excel_row, 14, data['overbreak_area'])
                updated_count += 1
    
    print(f"  更新了 {updated_count} 个桩号")
    
    # 第三步：重新计算体积：(当前面积 + 下一个面积) / 2 * 25
    print(f"  步骤3: 计算体积...")
    for excel_row in range(10, ws.max_row, 2):  # 面积行
        current_design_area = get_cell_value(ws, excel_row, 12) or 0  # L列
        current_overbreak_area = get_cell_value(ws, excel_row, 14) or 0  # N列
        
        # 获取下一个断面的面积
        next_excel_row = excel_row + 2
        if next_excel_row <= ws.max_row:
            next_design_area = get_cell_value(ws, next_excel_row, 12) or 0
            next_overbreak_area = get_cell_value(ws, next_excel_row, 14) or 0
            
            # 计算体积：(当前面积 + 下一个面积) / 2 * 25
            try:
                design_volume = (float(current_design_area) + float(next_design_area)) / 2 * 25
                overbreak_volume = (float(current_overbreak_area) + float(next_overbreak_area)) / 2 * 25
            except:
                design_volume = 0
                overbreak_volume = 0
            
            # 填入工程量行（面积行的下一行）
            # M列(13)：开挖体积，O列(15)：超挖体积
            set_cell_value(ws, excel_row + 1, 13, round(design_volume, 2))
            set_cell_value(ws, excel_row + 1, 15, round(overbreak_volume, 2))

# 保存结果
output_file = target_file.replace('.xlsx', f'_替换后_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
wb.save(output_file)
wb.close()

print(f"\n处理完成!")
print(f"输出文件: {output_file}")
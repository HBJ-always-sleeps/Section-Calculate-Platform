# -*- coding: utf-8 -*-
"""
测试两个Excel文件的数据结构
"""
import pandas as pd
import openpyxl
import sys

# 文件路径
target_file = r"D:\断面算量平台\测试文件\2026年2月月进度工程量表（提交版）.xlsx"
source_file = r"D:\断面算量平台\测试文件\内湾段分层图（全航道）_分类汇总_20260316_154717.xlsx"

print("=" * 80, flush=True)
print("分析源文件", flush=True)
print("=" * 80, flush=True)

# 读取源文件
try:
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    print(f"源文件工作表: {wb_source.sheetnames}", flush=True)
    
    for sheet_name in wb_source.sheetnames:
        print(f"\n--- 工作表: {sheet_name} ---", flush=True)
        df = pd.read_excel(source_file, sheet_name=sheet_name, header=0)
        print(f"形状: {df.shape}", flush=True)
        print(f"列名: {list(df.columns)}", flush=True)
        print(f"前5行:\n{df.head()}", flush=True)
    wb_source.close()
except Exception as e:
    print(f"错误: {e}", flush=True)

print("\n" + "=" * 80, flush=True)
print("分析目标文件", flush=True)
print("=" * 80, flush=True)

# 读取目标文件
try:
    wb_target = openpyxl.load_workbook(target_file, data_only=True)
    print(f"目标文件工作表: {wb_target.sheetnames}", flush=True)
    
    for sheet_name in wb_target.sheetnames[:5]:  # 只看前5个sheet
        print(f"\n--- 工作表: {sheet_name} ---", flush=True)
        ws = wb_target[sheet_name]
        
        # 打印前20行的内容
        for row in range(1, min(25, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                row_data.append(str(val)[:15] if val else "")
            print(f"行{row}: {row_data}", flush=True)
    wb_target.close()
except Exception as e:
    print(f"错误: {e}", flush=True)

print("\n分析完成!", flush=True)
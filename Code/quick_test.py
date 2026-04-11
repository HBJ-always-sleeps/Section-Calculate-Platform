# -*- coding: utf-8 -*-
import pandas as pd
import openpyxl

target_file = r"D:\断面算量平台\测试文件\2026年2月月进度工程量表（提交版）.xlsx"

print("读取目标文件...", flush=True)
wb = openpyxl.load_workbook(target_file, data_only=True, read_only=True)
print(f"工作表: {wb.sheetnames}", flush=True)

# 只读取第一个sheet的前20行
ws = wb[wb.sheetnames[0]]
print(f"\n第一个工作表: {wb.sheetnames[0]}", flush=True)
for row in range(1, 20):
    row_data = []
    for col in range(1, 18):
        val = ws.cell(row=row, column=col).value
        row_data.append(str(val)[:12] if val else "")
    print(f"R{row}: {row_data}", flush=True)

wb.close()

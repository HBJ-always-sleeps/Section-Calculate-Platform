#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
autolabel.py - 自动标注更新模块
功能：
1. 从Excel工程量表汇总开挖面积和超挖面积数据
2. 更新DXF断面图中的面积标注
3. 自验证更新结果

使用方法：
    from autolabel import AutoLabel
    
    # 方式1：完整流程
    autolabel = AutoLabel(excel_path, dxf_path)
    autolabel.run(output_path)
    
    # 方式2：分步执行
    autolabel = AutoLabel()
    data = autolabel.summarize_excel(excel_path)
    autolabel.update_dxf(dxf_path, data, output_path)
"""

import openpyxl
from openpyxl.utils import get_column_letter
import ezdxf
import os
import re
import datetime
from typing import List, Dict, Optional, Tuple


class AutoLabel:
    """自动标注更新器"""
    
    # 面积单位符号
    AREA_UNIT = '\u33A1'  # ㎡
    
    def __init__(self, excel_path: str = None, dxf_path: str = None):
        """
        初始化
        
        Args:
            excel_path: Excel工程量表路径
            dxf_path: DXF断面图路径
        """
        self.excel_path = excel_path
        self.dxf_path = dxf_path
        self.summary_data = []
        self.verify_result = None
        
    def get_cell_value(self, sheet, row: int, col: int):
        """获取单元格值"""
        cell = sheet.cell(row=row, column=col)
        return cell.value
    
    def summarize_excel(self, excel_path: str = None) -> List[Dict]:
        """
        汇总Excel数据
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            汇总数据列表: [{'桩号': str, '开挖面积': float, '超挖面积': float}, ...]
        """
        if excel_path:
            self.excel_path = excel_path
            
        if not os.path.exists(self.excel_path):
            raise FileNotFoundError(f"Excel文件不存在: {self.excel_path}")
        
        print(f"[汇总] 读取Excel: {self.excel_path}")
        
        # 加载工作簿
        wb = openpyxl.load_workbook(self.excel_path, data_only=True)
        sheet_names = wb.sheetnames
        print(f"[汇总] 发现 {len(sheet_names)} 个sheet")
        
        # 数据结构: {行号: {'桩号': str, '开挖面积': float, '超挖面积': float}}
        summary_data = {}
        
        # 列映射：L=12(开挖), N=14(超挖), D=4(备用开挖), F=6(备用超挖), B=2(桩号)
        COL_L = 12   # 开挖面积
        COL_N = 14   # 超挖面积
        COL_D = 4    # 备用开挖面积
        COL_F = 6    # 备用超挖面积
        COL_PILE = 2 # 桩号列
        
        # 遍历每个sheet
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            
            # 遍历行10到498
            for row in range(10, 499):
                # 获取桩号
                pile_number = self.get_cell_value(sheet, row, COL_PILE)
                
                # 获取L列和N列的值
                l_value = self.get_cell_value(sheet, row, COL_L)
                n_value = self.get_cell_value(sheet, row, COL_N)
                
                # 如果L列为空，使用D列
                if l_value is None or (isinstance(l_value, str) and l_value.strip() == ''):
                    l_value = self.get_cell_value(sheet, row, COL_D)
                
                # 如果N列为空，使用F列
                if n_value is None or (isinstance(n_value, str) and n_value.strip() == ''):
                    n_value = self.get_cell_value(sheet, row, COL_F)
                
                # 转换为数值
                try:
                    excavation = float(l_value) if l_value is not None and l_value != '' else 0.0
                except (ValueError, TypeError):
                    excavation = 0.0
                
                try:
                    overbreak = float(n_value) if n_value is not None and n_value != '' else 0.0
                except (ValueError, TypeError):
                    overbreak = 0.0
                
                # 如果这一行有数据，则汇总
                if excavation != 0.0 or overbreak != 0.0 or pile_number is not None:
                    if row not in summary_data:
                        summary_data[row] = {
                            '桩号': str(pile_number) if pile_number else f"行{row}",
                            '开挖面积': 0.0,
                            '超挖面积': 0.0
                        }
                    
                    summary_data[row]['开挖面积'] += excavation
                    summary_data[row]['超挖面积'] += overbreak
                    
                    # 更新桩号
                    if pile_number is not None and pile_number != '':
                        summary_data[row]['桩号'] = str(pile_number)
        
        wb.close()
        
        # 转换为列表并排序
        self.summary_data = [summary_data[row] for row in sorted(summary_data.keys())]
        
        # 打印汇总
        total_excavation = sum(d['开挖面积'] for d in self.summary_data)
        total_overbreak = sum(d['超挖面积'] for d in self.summary_data)
        
        print(f"[汇总] 共 {len(self.summary_data)} 行数据")
        print(f"[汇总] 开挖面积合计: {total_excavation:.2f}")
        print(f"[汇总] 超挖面积合计: {total_overbreak:.2f}")
        
        return self.summary_data
    
    def sort_pile_number(self, pile_text: str) -> int:
        """桩号排序键值"""
        match = re.search(r'K(\d+)\+(\d+)', str(pile_text))
        if match:
            return int(match.group(1)) * 1000 + int(match.group(2))
        return 0
    
    def update_dxf(self, dxf_path: str = None, data: List[Dict] = None, 
                   output_path: str = None, area_layer: str = '0', 
                   pile_layer: str = '桩号') -> Tuple[bool, str]:
        """
        更新DXF面积标注
        
        Args:
            dxf_path: DXF文件路径
            data: 汇总数据
            output_path: 输出路径
            area_layer: 面积标注图层
            pile_layer: 桩号图层
            
        Returns:
            (成功与否, 输出文件路径或错误信息)
        """
        if dxf_path:
            self.dxf_path = dxf_path
        if data:
            self.summary_data = data
            
        if not os.path.exists(self.dxf_path):
            return False, f"DXF文件不存在: {self.dxf_path}"
        
        if not self.summary_data:
            return False, "没有汇总数据，请先调用 summarize_excel()"
        
        # 默认输出路径
        if not output_path:
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            output_dir = os.path.dirname(self.dxf_path)
            output_path = os.path.join(output_dir, f"面积标注更新_{timestamp}.dxf")
        
        print(f"[更新] 读取DXF: {self.dxf_path}")
        
        try:
            # 读取DXF
            doc = ezdxf.readfile(self.dxf_path)
            msp = doc.modelspace()
            
            # 准备新数据
            new_data = []
            for d in self.summary_data:
                pile_text = str(d['桩号']).strip()
                design_remain = float(d['开挖面积'])
                overbreak = float(d['超挖面积'])
                new_data.append({
                    '桩号': pile_text,
                    '设计剩余面积': design_remain,
                    '超挖面积': overbreak
                })
            new_data = sorted(new_data, key=lambda x: self.sort_pile_number(x['桩号']))
            
            # 提取所有文本
            all_texts = []
            for entity in msp:
                if hasattr(entity, 'dxf') and entity.dxftype() == 'TEXT':
                    all_texts.append({
                        'entity': entity,
                        'text': entity.dxf.text,
                        'layer': entity.dxf.layer,
                        'x': entity.dxf.insert[0],
                        'y': entity.dxf.insert[1]
                    })
            
            piles = [t for t in all_texts if t['layer'] == pile_layer]
            area_texts = [t for t in all_texts if t['layer'] == area_layer]
            print(f"[更新] 桩号: {len(piles)}, 面积标注: {len(area_texts)}")
            
            # 解析面积标注
            descriptions = []  # 描述文本
            values = []        # 数值文本
            
            for t in area_texts:
                text = t['text']
                match = re.search(r'(\d+\.?\d*)', text)
                
                if '本期总剩余面积=' in text:
                    t['type'] = '总剩余'
                    descriptions.append(t)
                elif '本期设计剩余面积=' in text:
                    t['type'] = '设计'
                    descriptions.append(t)
                elif '本期超挖剩余面积=' in text:
                    t['type'] = '超挖'
                    descriptions.append(t)
                elif match and ('O' in text or self.AREA_UNIT in text):
                    t['value'] = float(match.group(1))
                    values.append(t)
            
            print(f"[更新] 描述文本: {len(descriptions)}, 数值文本: {len(values)}")
            
            # 为数值文本关联类型
            for val in values:
                best_desc = None
                best_x_dist = float('inf')
                
                for desc in descriptions:
                    y_diff = abs(val['y'] - desc['y'])
                    if y_diff < 3:  # 同一行
                        x_dist = val['x'] - desc['x']
                        if 0 < x_dist < 100 and x_dist < best_x_dist:
                            best_x_dist = x_dist
                            best_desc = desc
                
                if best_desc:
                    val['type'] = best_desc['type']
                    val['desc_x'] = best_desc['x']
            
            # 按Y坐标分组桩号
            pile_y_groups = {}
            y_tolerance = 30.0
            for pile in piles:
                y = pile['y']
                found_key = None
                for key in pile_y_groups.keys():
                    if abs(key - y) < y_tolerance:
                        found_key = key
                        break
                if found_key:
                    pile_y_groups[found_key].append(pile)
                else:
                    pile_y_groups[y] = [pile]
            
            # 按Y坐标分组数值文本
            value_y_groups = {}
            for val in values:
                y = val['y']
                found_key = None
                for key in value_y_groups.keys():
                    if abs(key - y) < 3:
                        found_key = key
                        break
                if found_key:
                    value_y_groups[found_key].append(val)
                else:
                    value_y_groups[y] = [val]
            
            # 建立对应关系并更新
            sorted_pile_y = sorted(pile_y_groups.keys(), reverse=True)
            sorted_value_y = sorted(value_y_groups.keys(), reverse=True)
            
            update_count = 0
            verify_data = []
            
            for pile_y in sorted_pile_y:
                piles_in_group = pile_y_groups[pile_y]
                
                # 找桩号上方的数值Y层
                above_value_y = [y for y in sorted_value_y if y > pile_y]
                if len(above_value_y) < 3:
                    continue
                
                # 取最近的3层
                closest_3_y = sorted(above_value_y, key=lambda y: y - pile_y)[:3]
                
                # 按X坐标分左右
                piles_by_x = {}
                for pile in piles_in_group:
                    x = pile['x']
                    found = False
                    for key in piles_by_x:
                        if abs(key - x) < 100:
                            piles_by_x[key].append(pile)
                            found = True
                            break
                    if not found:
                        piles_by_x[x] = [pile]
                
                # 更新每个桩号的数值
                for pile_x_key, pile_list in piles_by_x.items():
                    pile = pile_list[0]
                    pile_text = pile['text']
                    
                    # 找对应的数据
                    pile_data = None
                    for d in new_data:
                        if d['桩号'] == pile_text:
                            pile_data = d
                            break
                    
                    if not pile_data:
                        continue
                    
                    design = pile_data['设计剩余面积']
                    overbreak = pile_data['超挖面积']
                    total = round(design + overbreak, 2)
                    
                    # 找这个X位置的数值文本
                    related_values = []
                    for y in closest_3_y:
                        for val in value_y_groups[y]:
                            if pile['x'] - 30 <= val['x'] <= pile['x'] + 150:
                                if 'type' in val:
                                    related_values.append(val)
                    
                    if len(related_values) < 3:
                        continue
                    
                    # 更新
                    updated = {'总剩余': None, '设计': None, '超挖': None}
                    for val in related_values:
                        if 'type' not in val:
                            continue
                        entity = val['entity']
                        old_val = val['value']
                        
                        if val['type'] == '总剩余':
                            new_text = f"{total:.2f}{self.AREA_UNIT}"
                            entity.dxf.text = new_text
                            updated['总剩余'] = {'old': old_val, 'new': total}
                            update_count += 1
                        elif val['type'] == '设计':
                            new_text = f"{design:.2f}{self.AREA_UNIT}"
                            entity.dxf.text = new_text
                            updated['设计'] = {'old': old_val, 'new': design}
                            update_count += 1
                        elif val['type'] == '超挖':
                            new_text = f"{overbreak:.2f}{self.AREA_UNIT}"
                            entity.dxf.text = new_text
                            updated['超挖'] = {'old': old_val, 'new': overbreak}
                            update_count += 1
                    
                    verify_data.append({
                        '桩号': pile_text,
                        '期望': {'总': total, '设计': design, '超挖': overbreak},
                        '更新': updated
                    })
            
            # 保存
            doc.saveas(output_path)
            print(f"[更新] 更新了 {update_count} 个数值")
            
            # ===== 验证 =====
            print("[验证] 重新读取验证...")
            
            doc2 = ezdxf.readfile(output_path)
            msp2 = doc2.modelspace()
            
            verify_values = []
            for entity in msp2:
                if hasattr(entity, 'dxf') and entity.dxftype() == 'TEXT':
                    if entity.dxf.layer == area_layer:
                        text = entity.dxf.text
                        match = re.search(r'(\d+\.?\d*)', text)
                        if match and (self.AREA_UNIT in text or 'O' in text):
                            verify_values.append({
                                'value': float(match.group(1)),
                                'x': entity.dxf.insert[0],
                                'y': entity.dxf.insert[1]
                            })
            
            success = 0
            fail = 0
            for v in verify_data:
                upd = v['更新']
                if upd['总剩余'] and upd['设计'] and upd['超挖']:
                    success += 1
                else:
                    fail += 1
            
            print(f"[验证] 成功: {success}, 失败: {fail}")
            
            self.verify_result = {
                'success': success,
                'fail': fail,
                'total': len(verify_data),
                'details': verify_data[:10]  # 只保留前10条详情
            }
            
            if fail == 0 and success > 0:
                print(f"[完成] 输出文件: {output_path}")
                return True, output_path
            else:
                return False, f"验证失败: {fail} 个桩号未正确更新"
                
        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, f"处理错误: {str(e)}"
    
    def run(self, excel_path: str = None, dxf_path: str = None, 
            output_path: str = None, area_layer: str = '0', 
            pile_layer: str = '桩号') -> Tuple[bool, str]:
        """
        完整流程：汇总Excel + 更新DXF
        
        Args:
            excel_path: Excel文件路径
            dxf_path: DXF文件路径
            output_path: 输出路径
            area_layer: 面积标注图层
            pile_layer: 桩号图层
            
        Returns:
            (成功与否, 输出文件路径或错误信息)
        """
        print("=" * 60)
        print("AutoLabel - 自动标注更新")
        print("=" * 60)
        
        # 1. 汇总Excel
        try:
            data = self.summarize_excel(excel_path)
        except Exception as e:
            return False, f"Excel汇总失败: {str(e)}"
        
        # 2. 更新DXF
        success, result = self.update_dxf(dxf_path, data, output_path, area_layer, pile_layer)
        
        print("=" * 60)
        if success:
            print("[OK] 处理完成")
        else:
            print("[FAIL] 处理失败")
        print("=" * 60)
        
        return success, result
    
    def save_summary_excel(self, output_path: str, data: List[Dict] = None) -> str:
        """
        保存汇总结果到Excel
        
        Args:
            output_path: 输出路径
            data: 数据（可选，默认使用已汇总的数据）
            
        Returns:
            输出文件路径
        """
        if data is None:
            data = self.summary_data
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "汇总结果"
        
        # 写入表头
        ws['A1'] = "桩号"
        ws['B1'] = "开挖面积"
        ws['C1'] = "超挖面积"
        
        # 写入数据
        for idx, item in enumerate(data, start=2):
            ws[f'A{idx}'] = item['桩号']
            ws[f'B{idx}'] = item['开挖面积']
            ws[f'C{idx}'] = item['超挖面积']
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        wb.save(output_path)
        print(f"[保存] 汇总结果: {output_path}")
        
        return output_path


# 命令行入口
if __name__ == "__main__":
    import sys
    
    # 默认路径
    DEFAULT_EXCEL = r"D:\2026年3月月进度工程量表v3(1).xlsx"
    DEFAULT_DXf = r"D:\2026年3月月进度测量段面图北海港铁山港20万吨级航道工程（啄罗作业区至石头埠作业区段）施工Ⅰ标段.dxf"
    
    # 解析参数
    excel_path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_EXCEL
    dxf_path = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_DXf
    
    # 输出路径
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dxf = os.path.join(os.path.dirname(dxf_path), f"面积标注更新_{timestamp}.dxf")
    
    # 执行
    autolabel = AutoLabel()
    success, result = autolabel.run(
        excel_path=excel_path,
        dxf_path=dxf_path,
        output_path=output_dxf,
        area_layer='0',
        pile_layer='桩号'
    )
    
    if success:
        print(f"\n输出文件: {result}")
    else:
        print(f"\n错误: {result}")
        sys.exit(1)
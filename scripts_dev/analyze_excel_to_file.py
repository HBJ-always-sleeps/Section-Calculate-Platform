# -*- coding: utf-8 -*-
import openpyxl

excel_path = r'D:\tunnel_build\测试文件\2026年2月月进度工程量表（提交版）.xlsx'

wb = openpyxl.load_workbook(excel_path, read_only=True)

with open(r'D:\tunnel_build\excel_structure_analysis.txt', 'w', encoding='utf-8') as f:
    f.write("="*60 + "\n")
    f.write("Sheet名称列表:\n")
    f.write("="*60 + "\n")
    for i, name in enumerate(wb.sheetnames):
        f.write(f"{i+1}. {name}\n")
    f.write(f"\n共 {len(wb.sheetnames)} 个sheet\n")
    
    # 分析每个sheet的结构
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write("\n" + "="*60 + "\n")
        f.write(f"Sheet: {sheet_name}\n")
        f.write(f"行数: {ws.max_row}, 列数: {ws.max_column}\n")
        f.write("="*60 + "\n")
        
        # 打印前25行的内容
        for row in range(1, min(26, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    row_data.append(f"[{col}]{cell.value}")
            if row_data:
                f.write(f"行{row}: {' | '.join(str(d) for d in row_data)}\n")

wb.close()
print("Done! Output saved to D:\\tunnel_build\\excel_structure_analysis.txt")
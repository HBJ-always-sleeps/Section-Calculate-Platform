# -*- coding: utf-8 -*-
"""
分析Excel模板文件结构
"""
import openpyxl
import os

# Excel文件路径
excel_path = r'D:\tunnel_build\测试文件\2026年2月月进度工程量表（提交版）.xlsx'

print(f"正在加载: {excel_path}")
print(f"文件存在: {os.path.exists(excel_path)}")

try:
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    print("\n" + "="*60)
    print("Sheet名称列表:")
    print("="*60)
    for i, name in enumerate(wb.sheetnames):
        print(f"{i+1}. {name}")
    print(f"\n共 {len(wb.sheetnames)} 个sheet")
    
    # 分析每个sheet的结构
    for sheet_name in wb.sheetnames[:3]:  # 先分析前3个sheet
        ws = wb[sheet_name]
        print("\n" + "="*60)
        print(f"Sheet: {sheet_name}")
        print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
        print("="*60)
        
        # 打印前20行的内容
        for row in range(1, min(21, ws.max_row + 1)):
            row_data = []
            for col in range(1, min(15, ws.max_column + 1)):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    row_data.append(f"[{col}]{cell.value}")
            if row_data:
                print(f"行{row}: {' | '.join(row_data)}")
    
    wb.close()
    print("\n分析完成!")
    
except Exception as e:
    print(f"错误: {e}")
    import traceback
    traceback.print_exc()
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel数据汇总脚本
功能：
1. 读取所有sheet的L10-L498和N10-N498数据
2. 如果L或N列为空，则使用D列和F列数据
3. 按行汇总所有sheet的数据
4. 输出：桩号、开挖面积、超挖面积三列表格
"""

import openpyxl
from openpyxl.utils import get_column_letter
import os

def get_cell_value(sheet, row, col):
    """获取单元格值，处理合并单元格"""
    cell = sheet.cell(row=row, column=col)
    return cell.value

def process_excel_file(file_path):
    """处理Excel文件"""
    print(f"正在处理文件: {file_path}")
    
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path, data_only=True)
    
    # 获取所有sheet名称
    sheet_names = wb.sheetnames
    print(f"发现 {len(sheet_names)} 个sheet: {sheet_names}")
    
    # 初始化汇总数据
    # 数据结构: {行号: {'pile_number': 桩号, 'excavation': 开挖面积, 'overbreak': 超挖面积}}
    summary_data = {}
    
    # 列映射：L=12, N=14, D=4, F=6
    COL_L = 12  # 开挖面积
    COL_N = 14  # 超挖面积
    COL_D = 4   # 备用开挖面积
    COL_F = 6   # 备用超挖面积
    COL_PILE = 2  # B列，假设桩号在B列
    
    # 遍历每个sheet
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        print(f"\n处理Sheet: {sheet_name}")
        
        # 遍历行10到498
        for row in range(10, 499):
            # 获取桩号（假设在B列）
            pile_number = get_cell_value(sheet, row, COL_PILE)
            
            # 获取L列和N列的值
            l_value = get_cell_value(sheet, row, COL_L)
            n_value = get_cell_value(sheet, row, COL_N)
            
            # 如果L列为空，使用D列
            if l_value is None or (isinstance(l_value, str) and l_value.strip() == ''):
                l_value = get_cell_value(sheet, row, COL_D)
                if isinstance(l_value, str):
                    l_value = l_value.strip()
            
            # 如果N列为空，使用F列
            if n_value is None or (isinstance(n_value, str) and n_value.strip() == ''):
                n_value = get_cell_value(sheet, row, COL_F)
                if isinstance(n_value, str):
                    n_value = n_value.strip()
            
            # 转换为数值
            try:
                excavation = float(l_value) if l_value is not None and l_value != '' else 0.0
            except (ValueError, TypeError):
                excavation = 0.0
            
            try:
                overbreak = float(n_value) if n_value is not None and n_value != '' else 0.0
            except (ValueError, TypeError):
                overbreak = 0.0
            
            # 如果这一行有数据，则汇总
            if excavation != 0.0 or overbreak != 0.0 or pile_number is not None:
                if row not in summary_data:
                    summary_data[row] = {
                        'pile_number': pile_number,
                        'excavation': 0.0,
                        'overbreak': 0.0
                    }
                
                summary_data[row]['excavation'] += excavation
                summary_data[row]['overbreak'] += overbreak
                
                # 更新桩号（如果当前sheet有桩号）
                if pile_number is not None and pile_number != '':
                    summary_data[row]['pile_number'] = pile_number
    
    wb.close()
    
    # 转换为列表并排序
    result = []
    for row in sorted(summary_data.keys()):
        result.append({
            'row': row,
            'pile_number': summary_data[row]['pile_number'],
            'excavation': summary_data[row]['excavation'],
            'overbreak': summary_data[row]['overbreak']
        })
    
    return result

def save_to_excel(data, output_path):
    """保存结果到Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "汇总结果"
    
    # 写入表头
    ws['A1'] = "桩号"
    ws['B1'] = "开挖面积"
    ws['C1'] = "超挖面积"
    
    # 写入数据
    for idx, item in enumerate(data, start=2):
        ws[f'A{idx}'] = item['pile_number']
        ws[f'B{idx}'] = item['excavation']
        ws[f'C{idx}'] = item['overbreak']
    
    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    wb.save(output_path)
    print(f"\n结果已保存到: {output_path}")

def print_summary(data):
    """打印汇总结果"""
    print("\n" + "="*60)
    print("汇总结果")
    print("="*60)
    print(f"{'桩号':<20} {'开挖面积':>15} {'超挖面积':>15}")
    print("-"*60)
    
    total_excavation = 0.0
    total_overbreak = 0.0
    
    for item in data:
        pile = str(item['pile_number']) if item['pile_number'] is not None else f"行{item['row']}"
        excavation = item['excavation']
        overbreak = item['overbreak']
        
        total_excavation += excavation
        total_overbreak += overbreak
        
        print(f"{pile:<20} {excavation:>15.2f} {overbreak:>15.2f}")
    
    print("-"*60)
    print(f"{'合计':<20} {total_excavation:>15.2f} {total_overbreak:>15.2f}")
    print("="*60)
    print(f"共 {len(data)} 行数据")

def main():
    # 输入文件路径
    input_file = r"D:\2026年3月月进度工程量表v3.xlsx"
    output_file = r"D:\汇总结果.xlsx"
    
    # 检查文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 文件不存在: {input_file}")
        return
    
    # 处理数据
    data = process_excel_file(input_file)
    
    # 打印结果
    print_summary(data)
    
    # 保存到Excel
    save_to_excel(data, output_file)
    
    # 同时保存为CSV（方便查看）
    csv_file = r"D:\汇总结果.csv"
    import csv
    with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(['桩号', '开挖面积', '超挖面积'])
        for item in data:
            writer.writerow([
                item['pile_number'],
                item['excavation'],
                item['overbreak']
            ])
    print(f"CSV格式已保存到: {csv_file}")

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
数据迁移脚本：从autoclassify输出迁移到月进度工程量表模板

功能：
1. 读取autoclassify生成的Excel汇总文件
2. 按土类分类数据
3. 计算工程量（体积）
4. 输出符合月进度模板格式的Excel文件
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
import datetime
import re


# ==================== 配置 ====================

# 土类到sheet名的映射
SOIL_TYPE_SHEET_MAP = {
    '1级淤泥': '1级淤泥',
    '2级淤泥': '2级淤泥',
    '3级淤泥': '3级淤泥',
    '1级填土': '1级填土',
    '4级填土': '4级填土',
    '3级粘土': '3级粘土',
    '4级粘土': '4级粘土',
    '5级粘土': '5级粘土',
    '6级砂': '6级砂',
    '7级砂': '7级砂',
    '8级砂': '8级砂',
    '6级碎石': '6级碎石',
    '9级碎石': '9级碎石',
}

# 默认间距（米）
DEFAULT_INTERVAL = 25

# 表头模板
HEADER_TEMPLATE = {
    'row1': '北海港铁山港20万吨级航道工程（啄罗作业区至石头埠作业区段）施工1标段',
    'row3': '疏浚工程量计算表',
    'row5_业主': '业主单位：北海市路港建设投资开发有限公司',
    'row5_监理': '监理单位：广州华申建设工程管理有限公司',
    'row6_施工': '施工单位：中交广州航道局有限公司',
    'row6_日期': f'测量日期：{datetime.datetime.now().strftime("%Y年%m月%d日")}',
}


# ==================== 核心函数 ====================

def parse_station(station_str):
    """解析桩号字符串，返回数值用于排序"""
    if pd.isna(station_str):
        return 0
    # 提取数字部分，如 K67+400 -> 67400
    nums = re.findall(r'\d+', str(station_str))
    if len(nums) >= 2:
        return int(nums[0]) * 1000 + int(nums[1])
    elif len(nums) == 1:
        return int(nums[0])
    return 0


def read_autoclassify_output(file_path):
    """读取autoclassify输出的Excel文件
    
    返回:
        dict: {地层: DataFrame} 每个地层的数据
    """
    print(f"读取文件: {file_path}")
    
    # 尝试读取明细表
    try:
        df = pd.read_excel(file_path, sheet_name='明细表')
        print(f"  读取到明细表，共 {len(df)} 条记录")
    except:
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            print(f"  读取第一个sheet，共 {len(df)} 条记录")
        except Exception as e:
            print(f"  读取失败: {e}")
            return {}
    
    # 按地层分组
    result = {}
    if '地层' in df.columns:
        for strata in df['地层'].unique():
            strata_df = df[df['地层'] == strata].copy()
            strata_df['sort_key'] = strata_df['桩号'].apply(parse_station)
            strata_df = strata_df.sort_values('sort_key')
            result[strata] = strata_df
            print(f"  地层 '{strata}': {len(strata_df)} 条记录")
    
    return result


def calculate_volume(area_list, interval=DEFAULT_INTERVAL):
    """计算工程量（体积）
    
    使用梯形公式: V = (A1 + A2) / 2 * L
    
    参数:
        area_list: 面积列表
        interval: 断面间距
    
    返回:
        volume_list: 体积列表
    """
    volumes = []
    n = len(area_list)
    
    for i in range(n):
        if i < n - 1:
            # 梯形公式
            v = (area_list[i] + area_list[i + 1]) / 2 * interval
        else:
            # 最后一个断面，体积为0
            v = 0
        volumes.append(v)
    
    return volumes


def create_soil_sheet(ws, soil_type, data_df, start_station=None, end_station=None):
    """创建土类sheet
    
    参数:
        ws: openpyxl worksheet
        soil_type: 土类名称（如"1级淤泥"）
        data_df: 该土类的数据DataFrame
        start_station: 起始桩号（可选，用于筛选）
        end_station: 结束桩号（可选，用于筛选）
    """
    # 写入表头
    ws['A1'] = HEADER_TEMPLATE['row1']
    ws.merge_cells('A1:P1')
    
    ws['A3'] = HEADER_TEMPLATE['row3']
    ws.merge_cells('A3:P3')
    
    ws['A5'] = HEADER_TEMPLATE['row5_业主']
    ws.merge_cells('A5:H5')
    ws['I5'] = HEADER_TEMPLATE['row5_监理']
    ws.merge_cells('I5:P5')
    
    ws['A6'] = HEADER_TEMPLATE['row6_施工']
    ws.merge_cells('A6:H6')
    ws['I6'] = HEADER_TEMPLATE['row6_日期']
    ws.merge_cells('I6:P6')
    
    # 第7-9行：列标题
    headers_row7 = ['序号', '桩号', '间距(m)', 
                    f'{soil_type}设计工程量', '', '', f'{soil_type}上期剩余工程量', '', '', 
                    f'{soil_type}本期剩余工程量', '', '', '', '', '本期完成工程量(m³)']
    headers_row8 = ['', '', '', '开挖', '', '超挖', '', '', '开挖', '', '超挖', '', '开挖', '', '超挖', '']
    headers_row9 = ['', '', '', '面积(m²)', '工程量(m³)', '面积(m²)', '工程量(m³)', 
                    '面积(m²)', '工程量(m³)', '面积(m²)', '工程量(m³)', 
                    '面积(m²)', '体积(m³)', '面积(m²)', '体积(m³)', '']
    
    for col, val in enumerate(headers_row7, 1):
        ws.cell(row=7, column=col, value=val)
    for col, val in enumerate(headers_row8, 1):
        ws.cell(row=8, column=col, value=val)
    for col, val in enumerate(headers_row9, 1):
        ws.cell(row=9, column=col, value=val)
    
    # 合并单元格
    ws.merge_cells('A7:A9')
    ws.merge_cells('B7:B9')
    ws.merge_cells('C7:C9')
    ws.merge_cells('D7:E8')
    ws.merge_cells('F7:G8')
    ws.merge_cells('H7:I8')
    ws.merge_cells('J7:K8')
    ws.merge_cells('L7:M8')
    ws.merge_cells('N7:O8')
    ws.merge_cells('P7:P9')
    
    # 写入数据
    if data_df is None or len(data_df) == 0:
        print(f"  {soil_type}: 无数据")
        return
    
    # 提取面积数据
    excav_areas = data_df['设计面积'].tolist() if '设计面积' in data_df.columns else []
    over_areas = data_df['超挖面积'].tolist() if '超挖面积' in data_df.columns else []
    stations = data_df['桩号'].tolist() if '桩号' in data_df.columns else []
    
    # 计算体积
    excav_volumes = calculate_volume(excav_areas)
    over_volumes = calculate_volume(over_areas)
    
    # 写入数据行（每个桩号2行）
    row = 10
    for i, station in enumerate(stations):
        # 第1行：序号、桩号、面积
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=station)
        ws.cell(row=row, column=4, value=round(excav_areas[i], 2) if i < len(excav_areas) else 0)
        ws.cell(row=row, column=6, value=round(over_areas[i], 2) if i < len(over_areas) else 0)
        
        # 第2行：间距、体积
        row += 1
        ws.cell(row=row, column=3, value=DEFAULT_INTERVAL)
        ws.cell(row=row, column=5, value=round(excav_volumes[i], 2) if i < len(excav_volumes) else 0)
        ws.cell(row=row, column=7, value=round(over_volumes[i], 2) if i < len(over_volumes) else 0)
        
        row += 1
    
    print(f"  {soil_type}: 写入 {len(stations)} 个桩号，{row - 10} 行数据")


def create_output_excel(data_by_strata, output_path, template_path=None):
    """创建输出Excel文件
    
    参数:
        data_by_strata: dict, {地层: DataFrame}
        output_path: 输出文件路径
        template_path: 模板文件路径（可选）
    """
    print(f"\n创建输出文件: {output_path}")
    
    wb = openpyxl.Workbook()
    
    # 删除默认sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 获取所有需要处理的土类
    all_strata = set(data_by_strata.keys()) | set(SOIL_TYPE_SHEET_MAP.keys())
    
    # 按土类顺序创建sheet
    strata_order = [
        '1级淤泥', '2级淤泥', '3级淤泥',
        '1级填土', '5级粘土', '4级填土', '3级粘土', '4级粘土',
        '6级砂', '7级砂', '8级砂',
        '6级碎石', '9级碎石'
    ]
    
    for strata in strata_order:
        ws = wb.create_sheet(title=strata)
        data_df = data_by_strata.get(strata)
        create_soil_sheet(ws, strata, data_df)
    
    # 保存
    wb.save(output_path)
    print(f"\n输出文件已保存: {output_path}")


def migrate_autoclassify_to_template(input_path, output_path=None, template_path=None):
    """主函数：迁移autoclassify数据到模板格式
    
    参数:
        input_path: autoclassify输出的Excel文件路径
        output_path: 输出文件路径（可选）
        template_path: 模板文件路径（可选）
    """
    # 读取数据
    data_by_strata = read_autoclassify_output(input_path)
    
    if not data_by_strata:
        print("未找到有效数据")
        return None
    
    # 生成输出路径
    if output_path is None:
        base_dir = os.path.dirname(input_path)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        base_name = os.path.splitext(os.path.basename(input_path))[0]
        output_path = os.path.join(base_dir, f"{base_name}_月进度格式_{timestamp}.xlsx")
    
    # 创建输出文件
    create_output_excel(data_by_strata, output_path, template_path)
    
    return output_path


# ==================== 测试入口 ====================

if __name__ == "__main__":
    import sys
    
    if len(sys.argv) < 2:
        print("用法: python migrate_to_template.py <autoclassify输出文件.xlsx> [输出文件.xlsx]")
        print("\n示例:")
        print("  python migrate_to_template.py 分类汇总_20260319.xlsx")
        print("  python migrate_to_template.py 分类汇总_20260319.xlsx 月进度工程量表.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else None
    
    result = migrate_autoclassify_to_template(input_file, output_file)
    
    if result:
        print(f"\n✅ 转换成功！输出文件: {result}")
    else:
        print("\n❌ 转换失败")